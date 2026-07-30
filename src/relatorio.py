"""Montagem e formatação da pasta de trabalho.

Sete abas, na ordem de leitura de quem decide — não na ordem em que foi fácil calcular.
O critério de pronto é o do roadmap: abrir o arquivo e entender a carteira em trinta
segundos, sem ninguém explicando. Tudo aqui serve a isso.

Três regras de formatação que valem para o arquivo inteiro:
  · nenhuma célula com número cru sem rótulo;
  · percentual formatado como percentual, data como data;
  · toda aba abre com a pergunta de decisão que ela responde escrita no topo.
"""

from __future__ import annotations

import datetime as dt

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# --------------------------------------------------------------------------------------
# Paleta e tipografia. Um lugar só, para o arquivo inteiro parecer uma peça só.
# --------------------------------------------------------------------------------------

FONTE = "Segoe UI"

TINTA = "16324F"        # azul petróleo, cabeçalhos e títulos
GRAFITE = "3B4A5A"      # texto secundário
NEVOA = "8496A8"        # texto terciário, notas de rodapé
PAPEL = "FFFFFF"
CLARO = "F2F6FA"        # faixa alternada e fundo de cartão
LINHA = "DAE3EC"        # bordas

VERDE, VERDE_BG = "1B7F5A", "E4F3EC"
AMBAR, AMBAR_BG = "9A6B08", "FBF0DA"
VERMELHO, VERMELHO_BG = "A83232", "FAE6E6"
ACENTO = "2E6F9E"

CORES_GRAFICO = ["2E6F9E", "5FA8D3", "E8A33D", "A83232"]

_FINA = Side(style="thin", color=LINHA)
BORDA_TABELA = Border(bottom=_FINA)


def _fill(cor: str) -> PatternFill:
    return PatternFill("solid", fgColor=cor)


def _fonte(tamanho=10, cor=GRAFITE, negrito=False, italico=False) -> Font:
    return Font(name=FONTE, size=tamanho, color=cor, bold=negrito, italic=italico)


# --------------------------------------------------------------------------------------
# Blocos de construção
# --------------------------------------------------------------------------------------

def _preparar_aba(ws: Worksheet, cor_aba: str = ACENTO) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = cor_aba
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _cabecalho_aba(ws: Worksheet, titulo: str, pergunta: str, largura: int = 10) -> int:
    """Escreve título e pergunta de decisão. Devolve a linha em que a tabela começa."""
    ws.column_dimensions["A"].width = 2.2
    ws["B2"] = titulo
    ws["B2"].font = _fonte(16, TINTA, negrito=True)
    ws.row_dimensions[2].height = 24

    ws["B3"] = pergunta
    ws["B3"].font = _fonte(10, NEVOA, italico=True)
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=max(3, largura + 1))
    ws.row_dimensions[3].height = 16
    return 5


def _formatar_valor(valor):
    """Converte o que o Excel não sabe guardar; o resto passa direto."""
    if valor is None or (isinstance(valor, float) and np.isnan(valor)):
        return None
    if isinstance(valor, (np.integer,)):
        return int(valor)
    if isinstance(valor, (np.floating,)):
        return float(valor)
    if isinstance(valor, (np.bool_, bool)):
        return "Sim" if bool(valor) else "Não"
    if isinstance(valor, pd.Timestamp):
        return valor.to_pydatetime().date()
    return valor


def escrever_tabela(
    ws: Worksheet,
    df: pd.DataFrame,
    linha_inicial: int,
    rotulos: dict[str, str] | None = None,
    formatos: dict[str, str] | None = None,
    larguras: dict[str, float] | None = None,
    coluna_inicial: int = 2,
    banda: bool = True,
    interativo: bool = True,
) -> tuple[int, int]:
    """Despeja um DataFrame formatado. Devolve (primeira_linha_de_dados, ultima_linha)."""
    rotulos = rotulos or {}
    formatos = formatos or {}
    larguras = larguras or {}

    linha_cab = linha_inicial
    for j, coluna in enumerate(df.columns):
        cel = ws.cell(row=linha_cab, column=coluna_inicial + j)
        cel.value = rotulos.get(coluna, coluna.replace("_", " ").capitalize())
        cel.font = _fonte(10, PAPEL, negrito=True)
        cel.fill = _fill(TINTA)
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[linha_cab].height = 30

    for i, (_, linha) in enumerate(df.iterrows()):
        r = linha_cab + 1 + i
        fundo = _fill(CLARO) if (banda and i % 2 == 1) else None
        for j, coluna in enumerate(df.columns):
            cel = ws.cell(row=r, column=coluna_inicial + j)
            cel.value = _formatar_valor(linha[coluna])
            cel.font = _fonte(10, GRAFITE)
            cel.border = BORDA_TABELA
            if fundo:
                cel.fill = fundo
            if coluna in formatos:
                cel.number_format = formatos[coluna]
                cel.alignment = Alignment(horizontal="center")
            elif isinstance(cel.value, (int, float)):
                cel.number_format = "#,##0"
                cel.alignment = Alignment(horizontal="center")
            else:
                cel.alignment = Alignment(horizontal="left", vertical="center")

    ultima = linha_cab + len(df)

    for j, coluna in enumerate(df.columns):
        letra = get_column_letter(coluna_inicial + j)
        if coluna in larguras:
            ws.column_dimensions[letra].width = larguras[coluna]
        else:
            titulo = rotulos.get(coluna, coluna)
            maior = max(
                [len(str(titulo))]
                + [len(str(v)) for v in df[coluna].head(200).astype(str)]
            )
            ws.column_dimensions[letra].width = min(max(maior + 3, 11), 46)

    if interativo:
        ws.freeze_panes = ws.cell(row=linha_cab + 1, column=coluna_inicial)
        ws.auto_filter.ref = (
            f"{get_column_letter(coluna_inicial)}{linha_cab}:"
            f"{get_column_letter(coluna_inicial + len(df.columns) - 1)}{ultima}"
        )
        ws.print_title_rows = f"{linha_cab}:{linha_cab}"
    return linha_cab + 1, ultima


def _nota(ws: Worksheet, linha: int, texto: str, largura: int = 9) -> None:
    cel = ws.cell(row=linha, column=2)
    cel.value = texto
    cel.font = _fonte(9, NEVOA, italico=True)
    ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=2 + largura)
    cel.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[linha].height = 26


def _pintar_por_classificacao(
    ws: Worksheet, coluna: int, primeira: int, ultima: int, valores: pd.Series
) -> None:
    cores = {
        "Crítico": (VERMELHO, VERMELHO_BG),
        "Atenção": (AMBAR, AMBAR_BG),
        "Saudável": (VERDE, VERDE_BG),
        "Estável": (VERDE, VERDE_BG),
    }
    for i, valor in enumerate(valores):
        par = cores.get(str(valor))
        if not par:
            continue
        cel = ws.cell(row=primeira + i, column=coluna)
        cel.font = _fonte(10, par[0], negrito=True)
        cel.fill = _fill(par[1])
        cel.alignment = Alignment(horizontal="center")


# --------------------------------------------------------------------------------------
# Aba 1 · Resumo
# --------------------------------------------------------------------------------------

_LARGURA_CARTAO = 4
_COLUNAS_CARTAO = [2, 7, 12]   # B, G, L — com colunas-respiro entre eles


def _cartao(
    ws: Worksheet,
    linha: int,
    coluna: int,
    rotulo: str,
    valor: float,
    formato: str,
    contexto: str,
    pergunta: str,
    cor: str,
    cor_fundo: str,
) -> None:
    """Um indicador em bloco: rótulo, número grande, contexto e a pergunta que ele responde."""
    fim = coluna + _LARGURA_CARTAO - 1

    for r in range(linha, linha + 5):
        for c in range(coluna, fim + 1):
            cel = ws.cell(row=r, column=c)
            cel.fill = _fill(cor_fundo)
            cel.border = Border(
                top=Side(style="thin", color=cor) if r == linha else None,
                bottom=Side(style="thin", color=LINHA) if r == linha + 4 else None,
                left=Side(style="thin", color=LINHA) if c == coluna else None,
                right=Side(style="thin", color=LINHA) if c == fim else None,
            )

    # a faixa superior colorida é o único código de cor do painel: verde, âmbar, vermelho
    for c in range(coluna, fim + 1):
        ws.cell(row=linha, column=c).fill = _fill(cor)
    ws.row_dimensions[linha].height = 5

    ws.merge_cells(start_row=linha + 1, start_column=coluna, end_row=linha + 1, end_column=fim)
    cel = ws.cell(row=linha + 1, column=coluna)
    cel.value = rotulo.upper()
    cel.font = _fonte(9, NEVOA, negrito=True)
    cel.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[linha + 1].height = 20

    ws.merge_cells(start_row=linha + 2, start_column=coluna, end_row=linha + 2, end_column=fim)
    cel = ws.cell(row=linha + 2, column=coluna)
    # número guardado como número, com o formato fazendo o resto. Texto que parece número
    # faz o Excel desenhar o triângulo verde de erro no canto — e ele aparece no print.
    cel.value = valor
    cel.number_format = formato
    cel.font = Font(name=FONTE, size=26, bold=True, color=cor)
    cel.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[linha + 2].height = 34

    ws.merge_cells(start_row=linha + 3, start_column=coluna, end_row=linha + 3, end_column=fim)
    cel = ws.cell(row=linha + 3, column=coluna)
    cel.value = contexto
    cel.font = _fonte(9, GRAFITE)
    cel.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[linha + 3].height = 16

    ws.merge_cells(start_row=linha + 4, start_column=coluna, end_row=linha + 4, end_column=fim)
    cel = ws.cell(row=linha + 4, column=coluna)
    cel.value = pergunta
    cel.font = _fonte(9, NEVOA, italico=True)
    cel.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws.row_dimensions[linha + 4].height = 30


def _limpar_moldura(grafico) -> None:
    """Tira a caixa em volta do gráfico. Moldura em painel é ruído."""
    grafico.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    # a série da cobertura mora em colunas ocultas, e o Excel ignora célula oculta por
    # padrão — sem isto o gráfico sai em branco. O nome do atributo no openpyxl é este,
    # não `plotVisOnly`, que é como a propriedade se chama no arquivo gerado.
    grafico.visible_cells_only = False
    # openpyxl marca os eixos como apagados se não for dito o contrário, e o gráfico
    # perde os rótulos. Quem quiser esconder um eixo faz isso depois, caso a caso.
    grafico.x_axis.delete = False
    grafico.y_axis.delete = False


def _rotulos_so_valor() -> DataLabelList:
    rotulos = DataLabelList()
    rotulos.showVal = True
    rotulos.showSerName = False
    rotulos.showCatName = False
    rotulos.showLegendKey = False
    rotulos.showBubbleSize = False
    rotulos.showPercent = False
    return rotulos


def _agenda_compacta(ws: Worksheet, linha: int, ranking: pd.DataFrame, n: int = 8) -> int:
    """Os n clientes no topo do risco, em lista.

    Lista e não tabela: a tabela obrigaria colunas largas, e coluna larga aqui
    desmancharia a grade dos cartões acima — as larguras são as mesmas da planilha
    inteira. Em lista, cada linha cabe nos blocos que os cartões já usam.
    """
    cores = {"Crítico": VERMELHO, "Atenção": AMBAR, "Estável": VERDE}
    fundos = {"Crítico": VERMELHO_BG, "Atenção": AMBAR_BG, "Estável": VERDE_BG}

    for i, (_, cliente) in enumerate(ranking.head(n).iterrows()):
        r = linha + i
        fundo = _fill(CLARO) if i % 2 == 1 else None
        for c in range(2, 16):
            cel = ws.cell(row=r, column=c)
            if fundo:
                cel.fill = fundo
            cel.border = BORDA_TABELA

        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        cel = ws.cell(row=r, column=2)
        cel.value = f"{i + 1}.  {cliente['nome']}"
        cel.font = _fonte(10, TINTA, negrito=True)
        cel.alignment = Alignment(vertical="center", indent=1)

        cel = ws.cell(row=r, column=5)
        cel.value = float(cliente["score_risco"])
        cel.number_format = "0.0"
        cel.font = _fonte(11, cores.get(str(cliente["classificacao"]), GRAFITE), negrito=True)
        cel.alignment = Alignment(horizontal="right")

        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
        cel = ws.cell(row=r, column=7)
        cel.value = str(cliente["classificacao"])
        cel.font = _fonte(9, cores.get(str(cliente["classificacao"]), GRAFITE), negrito=True)
        cel.fill = _fill(fundos.get(str(cliente["classificacao"]), PAPEL))
        cel.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=15)
        cel = ws.cell(row=r, column=9)
        cel.value = (
            f"{int(cliente['pendencias_vencidas'])} pendências vencidas  ·  "
            f"aging médio de {cliente['aging_medio']:.0f} dias  ·  "
            f"{cliente['taxa_replanejamento']:.0%} das ações replanejadas"
        )
        cel.font = _fonte(9, GRAFITE)
        cel.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[r].height = 19

    return linha + n - 1


def _faixa_por_limite(valor: float, bom: float, ruim: float, maior_e_melhor: bool = True):
    if np.isnan(valor):
        return NEVOA, CLARO
    if maior_e_melhor:
        if valor >= bom:
            return VERDE, VERDE_BG
        return (AMBAR, AMBAR_BG) if valor >= ruim else (VERMELHO, VERMELHO_BG)
    if valor <= bom:
        return VERDE, VERDE_BG
    return (AMBAR, AMBAR_BG) if valor <= ruim else (VERMELHO, VERMELHO_BG)


def montar_resumo(
    ws: Worksheet,
    painel: dict,
    cobertura_semanal: pd.DataFrame,
    ranking: pd.DataFrame,
    aging: pd.DataFrame,
    ws_aging: Worksheet,
    cfg: dict,
    periodo: tuple[dt.date, dt.date],
) -> None:
    _preparar_aba(ws, TINTA)
    leitura = cfg["leitura"]

    for c in range(2, 16):
        ws.column_dimensions[get_column_letter(c)].width = 12.5
    ws.column_dimensions["A"].width = 2.2
    for c in (6, 11):  # colunas-respiro entre os cartões
        ws.column_dimensions[get_column_letter(c)].width = 2.0

    # ---- cabeçalho -------------------------------------------------------------------
    # faixa escura de ponta a ponta: é o que faz a captura de tela parecer um painel,
    # e não uma planilha com texto em cima
    for r in range(1, 6):
        for c in range(1, 17):
            ws.cell(row=r, column=c).fill = _fill(TINTA)
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[5].height = 10

    ws.merge_cells("B2:O2")
    ws["B2"] = "Painel de aderência a plano de ação"
    ws["B2"].font = Font(name=FONTE, size=22, bold=True, color=PAPEL)
    ws["B2"].alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 32

    inicio, fim = periodo
    ws.merge_cells("B3:O3")
    ws["B3"] = (
        f"Cenário {cfg['cenario']['nome']} · {cfg['cenario']['n_clientes']} clientes · "
        f"período de {inicio.strftime('%d/%m/%Y')} a {fim.strftime('%d/%m/%Y')} "
        f"({cfg['cenario']['horizonte_semanas']} semanas) · "
        f"{painel['total_acoes']} ações acompanhadas"
    )
    ws["B3"].font = _fonte(10, "C7D6E4")
    ws.row_dimensions[3].height = 18

    ws.merge_cells("B4:O4")
    ws["B4"] = "Dados sintéticos, gerados por código com semente fixa. Nenhum dado de cliente real."
    ws["B4"].font = _fonte(9, "8DA6BF", italico=True)
    ws.row_dimensions[6].height = 8

    # ---- seis cartões ---------------------------------------------------------------
    ader = painel["aderencia"]
    cor_ader = _faixa_por_limite(ader, leitura["aderencia_saudavel"], leitura["aderencia_critica"])
    pct30 = painel["aging_mais_30_pct"]
    cor30 = _faixa_por_limite(pct30, 0.15, 0.30, maior_e_melhor=False)
    cor_repl = _faixa_por_limite(
        painel["taxa_replanejamento"], 0.25, 0.45, maior_e_melhor=False
    )
    cor_cob = _faixa_por_limite(painel["cobertura"], leitura["cobertura_minima"], 0.55)
    cor_risco = (
        (VERMELHO, VERMELHO_BG) if painel["clientes_criticos"] >= 5
        else (AMBAR, AMBAR_BG) if painel["clientes_criticos"] >= 2
        else (VERDE, VERDE_BG)
    )

    cartoes = [
        (
            "Aderência ao prazo",
            ader,
            "0%",
            f"{painel['aderencia_denominador']} ações com prazo vencido no período",
            "O acompanhamento está funcionando ou só registrando?",
            *cor_ader,
        ),
        (
            "Pendências +30 dias",
            painel["aging_mais_30"],
            "0",
            f"de {painel['pendencias_abertas']} pendências em aberto ({pct30:.0%})",
            "Onde intervir primeiro. +30 dias não é atraso, é ação abandonada.",
            *cor30,
        ),
        (
            "Taxa de replanejamento",
            painel["taxa_replanejamento"],
            "0%",
            "das ações tiveram o prazo movido ao menos uma vez",
            "O prazo é mal estimado na origem ou a execução está travando?",
            *cor_repl,
        ),
        (
            "Lead time mediano",
            painel["lead_time_mediano"],
            '0" dias"',
            "da abertura à conclusão, mediana da carteira",
            "Quanto tempo prometer na próxima reunião, com base em histórico.",
            ACENTO,
            CLARO,
        ),
        (
            "Cobertura de acompanhamento",
            painel["cobertura"],
            "0%",
            "das semanas-cliente tiveram reunião registrada",
            (
                "Cobertura abaixo do mínimo: os números acima descrevem a parte acompanhada "
                "da carteira, não a carteira."
                if not painel["cobertura_suficiente"]
                else "Cobertura suficiente para ler os indicadores acima como carteira."
            ),
            *cor_cob,
        ),
        (
            "Clientes em risco crítico",
            painel["clientes_criticos"],
            "0",
            f"topo do ranking: {painel['cliente_topo_risco']}",
            "Por quem começar a agenda da semana.",
            *cor_risco,
        ),
    ]

    for i, cartao in enumerate(cartoes):
        linha = 7 + (i // 3) * 6
        _cartao(ws, linha, _COLUNAS_CARTAO[i % 3], *cartao)

    # ---- gráficos -------------------------------------------------------------------
    ws["B19"] = "Onde estão as pendências em aberto"
    ws["B19"].font = _fonte(12, TINTA, negrito=True)
    ws["I19"] = "Cobertura de acompanhamento, semana a semana"
    ws["I19"].font = _fonte(12, TINTA, negrito=True)
    ws.row_dimensions[19].height = 22

    grafico_aging = BarChart()
    grafico_aging.type = "col"
    grafico_aging.style = None
    grafico_aging.title = None
    grafico_aging.height, grafico_aging.width = 6.4, 11.4
    dados = Reference(ws_aging, min_col=3, min_row=5, max_row=5 + len(aging))
    cats = Reference(ws_aging, min_col=2, min_row=6, max_row=5 + len(aging))
    grafico_aging.add_data(dados, titles_from_data=True)
    grafico_aging.set_categories(cats)
    grafico_aging.legend = None
    grafico_aging.y_axis.majorGridlines = None
    grafico_aging.y_axis.delete = True
    grafico_aging.gapWidth = 60
    grafico_aging.dLbls = _rotulos_so_valor()
    _limpar_moldura(grafico_aging)
    grafico_aging.y_axis.delete = True  # o rótulo em cima da barra já diz o número

    serie = grafico_aging.series[0]
    serie.graphicalProperties.line.noFill = True
    # cada faixa com a cor da decisão que ela pede: as duas primeiras são rotina,
    # a terceira é cobrança, a quarta é ação abandonada
    cores_faixa = [ACENTO, ACENTO, AMBAR, VERMELHO]
    serie.data_points = [
        DataPoint(idx=i, spPr=GraphicalProperties(solidFill=cor))
        for i, cor in enumerate(cores_faixa[: len(aging)])
    ]
    ws.add_chart(grafico_aging, "B20")

    # série da cobertura em colunas auxiliares, escondidas: gráfico precisa de células
    col_aux = 20  # T
    ws.cell(row=1, column=col_aux, value="semana")
    ws.cell(row=1, column=col_aux + 1, value="cobertura")
    for i, (_, linha_cob) in enumerate(cobertura_semanal.iterrows(), start=2):
        ws.cell(row=i, column=col_aux, value=int(linha_cob["semana"]))
        ws.cell(row=i, column=col_aux + 1, value=float(linha_cob["cobertura"]))
    for c in (col_aux, col_aux + 1):
        ws.column_dimensions[get_column_letter(c)].hidden = True

    n = len(cobertura_semanal)
    grafico_cob = LineChart()
    grafico_cob.height, grafico_cob.width = 6.4, 11.4
    dados = Reference(ws, min_col=col_aux + 1, min_row=1, max_row=n + 1)
    grafico_cob.add_data(dados, titles_from_data=True)
    grafico_cob.set_categories(Reference(ws, min_col=col_aux, min_row=2, max_row=n + 1))
    grafico_cob.legend = None
    grafico_cob.y_axis.numFmt = "0%"
    grafico_cob.y_axis.scaling.min = 0
    grafico_cob.y_axis.scaling.max = 1
    grafico_cob.y_axis.majorGridlines.spPr = GraphicalProperties(
        ln=LineProperties(solidFill=LINHA, w=6350)
    )
    grafico_cob.x_axis.tickLblSkip = 4   # 26 semanas nomeadas uma a uma viram borrão
    grafico_cob.x_axis.tickMarkSkip = 4
    grafico_cob.series[0].graphicalProperties.line.solidFill = ACENTO
    grafico_cob.series[0].graphicalProperties.line.width = 22000
    grafico_cob.series[0].smooth = False
    _limpar_moldura(grafico_cob)
    ws.add_chart(grafico_cob, "I20")

    # ---- agenda da semana -----------------------------------------------------------
    ws["B34"] = "A agenda da semana"
    ws["B34"].font = _fonte(12, TINTA, negrito=True)
    ws.row_dimensions[34].height = 22
    ws["B35"] = "Os oito clientes no topo do ranking de risco, na ordem em que devem ser tratados."
    ws["B35"].font = _fonte(9, NEVOA, italico=True)

    ultima = _agenda_compacta(ws, 37, ranking, n=8)

    _nota(
        ws,
        ultima + 2,
        "O score é relativo à própria carteira: serve para ordenar a agenda, não para "
        "afirmar risco absoluto. As colunas cruas ao lado são a leitura honesta. "
        "Limitações completas em docs/limitacoes.md.",
        largura=12,
    )


# --------------------------------------------------------------------------------------
# Abas 2 a 7
# --------------------------------------------------------------------------------------

def montar_aderencia(ws: Worksheet, df: pd.DataFrame) -> None:
    _preparar_aba(ws)
    inicio = _cabecalho_aba(
        ws,
        "Aderência ao prazo, por cliente",
        "O acompanhamento está funcionando ou só registrando? — concluídas até o prazo "
        "ORIGINAL ÷ ações com prazo original vencido. Da pior para a melhor.",
        largura=8,
    )
    tabela = df[
        ["nome", "segmento", "porte", "consultor", "acoes_com_prazo_vencido",
         "concluidas_no_prazo", "pendencias_vencidas", "aderencia", "classificacao"]
    ]
    primeira, ultima = escrever_tabela(
        ws,
        tabela,
        inicio,
        rotulos={
            "nome": "Cliente",
            "segmento": "Segmento",
            "porte": "Porte",
            "consultor": "Consultor",
            "acoes_com_prazo_vencido": "Ações com prazo vencido",
            "concluidas_no_prazo": "Concluídas no prazo",
            "pendencias_vencidas": "Ainda em aberto",
            "aderencia": "Aderência",
            "classificacao": "Situação",
        },
        formatos={"aderencia": "0%"},
        larguras={"nome": 30},
    )
    col_ader = 9
    ws.conditional_formatting.add(
        f"{get_column_letter(col_ader)}{primeira}:{get_column_letter(col_ader)}{ultima}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=ACENTO),
    )
    _pintar_por_classificacao(ws, 10, primeira, ultima, tabela["classificacao"])
    _nota(
        ws,
        ultima + 2,
        "Ações canceladas ficam fora do cálculo. Uma ação vencida e ainda em aberto conta "
        "como descumprimento no período — esperar a conclusão para contabilizar seria adiar "
        "a má notícia.",
    )


def montar_aging(ws: Worksheet, faixas: pd.DataFrame, detalhe: pd.DataFrame) -> None:
    _preparar_aba(ws)
    inicio = _cabecalho_aba(
        ws,
        "Aging de pendências",
        "Onde intervir primeiro? — dias em aberto das pendências, por faixa. "
        "Pendência de +30 dias não é atraso, é ação abandonada.",
        largura=6,
    )
    primeira, ultima = escrever_tabela(
        ws,
        faixas,
        inicio,
        rotulos={
            "faixa": "Faixa",
            "pendencias": "Pendências",
            "percentual": "% do aberto",
            "dias_em_aberto_medio": "Dias em aberto (média)",
            "das_quais_prioridade_alta": "Das quais prioridade Alta",
            "leitura": "Decisão que a faixa pede",
        },
        formatos={"percentual": "0%", "dias_em_aberto_medio": "0.0"},
        larguras={"faixa": 14, "leitura": 42},
        banda=False,
    )
    # a faixa +30 é a única linha destacada do relatório inteiro; é o que se quer que
    # os olhos encontrem primeiro
    for c in range(2, 8):
        cel = ws.cell(row=ultima, column=c)
        cel.fill = _fill(VERMELHO_BG)
        cel.font = _fonte(10, VERMELHO, negrito=True)

    linha_detalhe = ultima + 3
    ws.cell(row=linha_detalhe, column=2, value="Pendências em aberto, da mais antiga para a mais recente")
    ws.cell(row=linha_detalhe, column=2).font = _fonte(12, TINTA, negrito=True)

    inicio_det = linha_detalhe + 2
    primeira_det, ultima_det = escrever_tabela(
        ws,
        detalhe,
        inicio_det,
        rotulos={
            "cliente": "Cliente",
            "acao_id": "Ação",
            "categoria": "Categoria",
            "descricao": "Descrição",
            "responsavel": "Responsável",
            "prioridade": "Prioridade",
            "data_abertura": "Aberta em",
            "prazo_original": "Prazo original",
            "prazo_atual": "Prazo atual",
            "n_reprogramacoes": "Reprogr.",
            "dias_em_aberto": "Dias em aberto",
            "dias_apos_prazo_original": "Dias após o prazo original",
            "faixa_aging": "Faixa",
            "status": "Status",
        },
        formatos={
            "data_abertura": "dd/mm/yyyy",
            "prazo_original": "dd/mm/yyyy",
            "prazo_atual": "dd/mm/yyyy",
            "dias_em_aberto": "0",
            "dias_apos_prazo_original": "0",
        },
        larguras={"cliente": 26, "descricao": 46},
    )
    ws.freeze_panes = ws.cell(row=inicio_det + 1, column=2)
    col_dias = 2 + list(detalhe.columns).index("dias_em_aberto")
    ws.conditional_formatting.add(
        f"{get_column_letter(col_dias)}{primeira_det}:{get_column_letter(col_dias)}{ultima_det}",
        ColorScaleRule(
            start_type="min", start_color=VERDE_BG,
            mid_type="percentile", mid_value=50, mid_color=AMBAR_BG,
            end_type="max", end_color=VERMELHO_BG,
        ),
    )
    _nota(
        ws,
        ultima_det + 2,
        "O aging conta a partir da abertura, não do vencimento: o que interessa é há quanto "
        "tempo o assunto está na pauta sem sair dela. A coluna de dias após o prazo original "
        "está ao lado para quem quiser a outra leitura.",
        largura=13,
    )


def montar_replanejamento(
    ws: Worksheet, por_cliente: pd.DataFrame, por_categoria: pd.DataFrame
) -> None:
    _preparar_aba(ws)
    inicio = _cabecalho_aba(
        ws,
        "Taxa de replanejamento",
        "O prazo está sendo mal estimado na origem ou a execução está travando? — "
        "ações com ao menos uma reprogramação ÷ ações não canceladas.",
        largura=7,
    )
    ws.cell(row=inicio, column=2, value="Por cliente").font = _fonte(12, TINTA, negrito=True)
    primeira, ultima = escrever_tabela(
        ws,
        por_cliente,
        inicio + 1,
        rotulos={
            "cliente_id": "ID",
            "nome": "Cliente",
            "acoes": "Ações",
            "acoes_replanejadas": "Replanejadas",
            "taxa_replanejamento": "Taxa",
            "reprogramacoes_totais": "Reprogramações",
            "repro_por_acao_replanejada": "Reprogr. por ação replanejada",
        },
        formatos={"taxa_replanejamento": "0%", "repro_por_acao_replanejada": "0.00"},
        larguras={"nome": 30, "cliente_id": 8},
    )
    col_taxa = 6
    ws.conditional_formatting.add(
        f"{get_column_letter(col_taxa)}{primeira}:{get_column_letter(col_taxa)}{ultima}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=AMBAR),
    )

    linha_cat = ultima + 3
    ws.cell(row=linha_cat, column=2, value="Por categoria de ação").font = _fonte(12, TINTA, negrito=True)
    escrever_tabela(
        ws,
        por_categoria,
        linha_cat + 1,
        rotulos={
            "categoria": "Categoria",
            "acoes": "Ações",
            "acoes_replanejadas": "Replanejadas",
            "reprogramacoes_totais": "Reprogramações",
            "taxa_replanejamento": "Taxa",
        },
        formatos={"taxa_replanejamento": "0%"},
        larguras={"categoria": 18},
    )
    ws.freeze_panes = ws.cell(row=inicio + 2, column=2)
    ws.auto_filter.ref = None

    _nota(
        ws,
        linha_cat + len(por_categoria) + 3,
        "Reprogramação por ação replanejada separa dois problemas diferentes: muitas ações "
        "movidas uma vez é erro de estimativa; poucas ações movidas quatro vezes é execução "
        "travada. Cruze com a aba Lead time para decidir qual dos dois é o seu caso.",
        largura=6,
    )


def montar_lead_time(ws: Worksheet, df: pd.DataFrame) -> None:
    _preparar_aba(ws)
    inicio = _cabecalho_aba(
        ws,
        "Lead time de conclusão",
        "Quanto tempo prometer na próxima reunião? — dias entre abertura e conclusão, "
        "por categoria. Mediana, não média: a distribuição tem cauda.",
        largura=8,
    )
    primeira, ultima = escrever_tabela(
        ws,
        df,
        inicio,
        rotulos={
            "categoria": "Categoria",
            "acoes_concluidas": "Ações concluídas",
            "p25": "p25",
            "mediana": "Mediana",
            "p75": "p75",
            "maximo": "Máximo",
            "amplitude_p25_p75": "Amplitude p25–p75",
            "mediana_com_pendentes": "Mediana incluindo pendentes",
            "prazo_tipico_prometido": "Prazo típico prometido",
            "diferenca_mediana_menos_prometido": "Diferença vs. prometido",
        },
        formatos={
            # uma casa decimal: mediana de contagem par cai no meio, e arredondar
            # para inteiro faria a subtração exibida não fechar com as parcelas
            "p25": "0.0", "mediana": "0.0", "p75": "0.0", "maximo": "0",
            "amplitude_p25_p75": "0.0", "mediana_com_pendentes": "0.0",
            "prazo_tipico_prometido": "0.0",
            "diferenca_mediana_menos_prometido": "+0.0;−0.0;0.0",
        },
        larguras={"categoria": 18},
        banda=False,
    )
    col_mediana = 2 + list(df.columns).index("mediana_com_pendentes")
    ws.conditional_formatting.add(
        f"{get_column_letter(col_mediana)}{primeira}:{get_column_letter(col_mediana)}{ultima}",
        DataBarRule(start_type="min", end_type="max", color=ACENTO),
    )
    col_dif = 2 + list(df.columns).index("diferenca_mediana_menos_prometido")
    for i in range(len(df)):
        cel = ws.cell(row=primeira + i, column=col_dif)
        if isinstance(cel.value, (int, float)) and cel.value is not None:
            cor = VERMELHO if cel.value > 0 else VERDE
            cel.font = _fonte(10, cor, negrito=True)

    _nota(
        ws,
        ultima + 2,
        "Leia as duas medianas juntas, e prometa prazo pela MAIOR delas. A primeira só "
        "enxerga ações que fecharam — e as que mais demoram são justamente as que ainda não "
        "fecharam, então ela é otimista por construção. A segunda conta cada pendência pelo "
        "tempo já decorrido, que ainda vai crescer: é um piso garantido, mas fica baixa quando "
        "a categoria tem muita pendência recente. Nenhuma das duas é o número verdadeiro; as "
        "duas são pisos, e o verdadeiro está acima de ambas.",
        largura=10,
    )


def montar_ranking(ws: Worksheet, df: pd.DataFrame, cfg_risco: dict) -> None:
    _preparar_aba(ws)
    inicio = _cabecalho_aba(
        ws,
        "Ranking de risco por cliente",
        f"Por quem começar a agenda da semana? — score = "
        f"{cfg_risco['peso_vencidas']:.0%} pendências vencidas + "
        f"{cfg_risco['peso_aging']:.0%} aging médio + "
        f"{cfg_risco['peso_replanejamento']:.0%} replanejamento, normalizados na carteira.",
        largura=10,
    )
    primeira, ultima = escrever_tabela(
        ws,
        df,
        inicio,
        rotulos={
            "cliente_id": "ID",
            "nome": "Cliente",
            "consultor": "Consultor",
            "score_risco": "Score",
            "classificacao": "Situação",
            "pendencias_vencidas": "Pendências vencidas",
            "pct_vencidas": "% das ações",
            "aging_medio": "Aging médio (dias)",
            "taxa_replanejamento": "Replanejamento",
            "pendencias_abertas": "Pendências abertas",
            "acoes": "Ações no período",
        },
        formatos={
            "score_risco": "0.0", "pct_vencidas": "0%",
            "aging_medio": "0.0", "taxa_replanejamento": "0%",
        },
        larguras={"nome": 30, "cliente_id": 8, "consultor": 14},
    )
    ws.conditional_formatting.add(
        f"E{primeira}:E{ultima}",
        ColorScaleRule(
            start_type="min", start_color=VERDE_BG,
            mid_type="percentile", mid_value=50, mid_color=AMBAR_BG,
            end_type="max", end_color=VERMELHO_BG,
        ),
    )
    _pintar_por_classificacao(ws, 6, primeira, ultima, df["classificacao"])
    _nota(
        ws,
        ultima + 2,
        "Score relativo à carteira: o pior cliente tende a 100 mesmo estando bem em termos "
        "absolutos. Ele ordena a agenda; quem descreve a situação são as colunas cruas ao lado.",
        largura=10,
    )


def montar_base(ws: Worksheet, df: pd.DataFrame) -> None:
    _preparar_aba(ws, NEVOA)
    inicio = _cabecalho_aba(
        ws,
        "Base",
        "Tabela achatada, uma linha por ação, pronta para tabela dinâmica ou Power BI. "
        "É a fonte de tudo o que as outras abas mostram.",
        largura=20,
    )
    escrever_tabela(
        ws,
        df,
        inicio,
        rotulos={c: c.replace("_", " ").capitalize() for c in df.columns},
        formatos={
            "data_abertura": "dd/mm/yyyy",
            "prazo_original": "dd/mm/yyyy",
            "prazo_atual": "dd/mm/yyyy",
            "data_conclusao": "dd/mm/yyyy",
            "data_referencia": "dd/mm/yyyy",
            "lead_time": "0",
            "dias_em_aberto": "0",
            "dias_apos_prazo_original": "0",
        },
        larguras={"descricao": 46, "cliente": 26},
        banda=False,
    )


# --------------------------------------------------------------------------------------
# Orquestração
# --------------------------------------------------------------------------------------

def montar_workbook(res: dict, cfg: dict, periodo: tuple[dt.date, dt.date]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    # O critério de pronto do projeto diz que a semente fixa reproduz o mesmo resultado.
    # Sem isto o arquivo muda a cada execução, porque o Excel carimba a hora da geração
    # dentro do pacote — dois relatórios idênticos sairiam com hashes diferentes. A data
    # do documento passa a ser a data de referência do relatório, que é a informativa.
    carimbo = dt.datetime.combine(cfg["cenario"]["data_corte"], dt.time(0, 0))
    wb.properties.creator = "painel-aderencia-plano-acao"
    wb.properties.title = "Painel de aderência a plano de ação"
    wb.properties.description = (
        f"Cenário {cfg['cenario']['nome']}, semente {cfg['cenario']['semente']}. "
        "Dados sintéticos."
    )
    wb.properties.created = carimbo
    wb.properties.modified = carimbo

    ws_resumo = wb.create_sheet("Resumo")
    ws_ader = wb.create_sheet("Aderência por cliente")
    ws_aging = wb.create_sheet("Aging")
    ws_repl = wb.create_sheet("Replanejamento")
    ws_lead = wb.create_sheet("Lead time")
    ws_rank = wb.create_sheet("Ranking de risco")
    ws_base = wb.create_sheet("Base")

    # o Aging é montado antes do Resumo porque o gráfico do Resumo referencia as
    # células desta aba — dado em um lugar só, sem cópia
    montar_aging(ws_aging, res["aging_faixas"], res["aging_detalhe"])
    montar_resumo(
        ws_resumo, res["painel"], res["cobertura_semanal"], res["ranking"],
        res["aging_faixas"], ws_aging, cfg, periodo,
    )
    montar_aderencia(ws_ader, res["aderencia_cliente"])
    montar_replanejamento(ws_repl, res["replanejamento_cliente"], res["replanejamento_categoria"])
    montar_lead_time(ws_lead, res["lead_time"])
    montar_ranking(ws_rank, res["ranking"], cfg["risco"])
    montar_base(ws_base, res["base_achatada"])

    wb.active = 0
    return wb


def salvar_reproduzivel(wb: Workbook, destino, carimbo: dt.date) -> None:
    """Grava a pasta de trabalho de forma que a mesma semente produza o mesmo arquivo.

    Um .xlsx é um zip, e o zip guarda a hora de gravação de cada item dentro dele. Duas
    execuções idênticas sairiam com hashes diferentes só por causa disso — o que
    transformaria o critério "a semente fixa reproduz o mesmo resultado" em algo que só
    dá para verificar no olho. Aqui o arquivo é reescrito com a data de referência do
    relatório em todos os itens, e a verificação passa a ser uma comparação de hash.
    """
    import io
    import re
    import zipfile

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    data_fixa = (carimbo.year, carimbo.month, carimbo.day, 0, 0, 0)
    iso = f"{carimbo.isoformat()}T00:00:00Z"

    with zipfile.ZipFile(buffer) as origem:
        itens = sorted(origem.infolist(), key=lambda i: i.filename)
        with zipfile.ZipFile(destino, "w", zipfile.ZIP_DEFLATED) as saida:
            for item in itens:
                conteudo = origem.read(item.filename)
                if item.filename == "docProps/core.xml":
                    # o openpyxl sobrescreve a data de modificação com a hora da gravação
                    # no momento do save, depois do que foi definido em montar_workbook
                    conteudo = re.sub(
                        rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                        rb"\g<1>" + iso.encode() + rb"\g<2>",
                        conteudo,
                    )
                novo = zipfile.ZipInfo(item.filename, date_time=data_fixa)
                novo.compress_type = zipfile.ZIP_DEFLATED
                novo.external_attr = item.external_attr
                saida.writestr(novo, conteudo)
